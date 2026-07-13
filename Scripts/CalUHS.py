import argparse
from openpyxl import load_workbook


def to_float_or_zero(value):
    """Convert Excel cell value to float. Treat empty/non-numeric values as 0."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def add_uhs_to_sheet(
    input_path: str,
    output_path: str,
    sheet_name: str = "all_windows",
    alpha: float = 0.5,
):
    wb = load_workbook(input_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Read header row
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header is not None:
            headers[str(header).strip()] = col

    required_cols = [
        "learned_novel_target_recall",
        "novel_context_old_target_acc",
    ]

    missing = [c for c in required_cols if c not in headers]
    if missing:
        raise ValueError(f"Missing required columns in sheet '{sheet_name}': {missing}")

    recall_col = headers["learned_novel_target_recall"]
    context_acc_col = headers["novel_context_old_target_acc"]

    # If UHS column already exists, overwrite it; otherwise append a new column.
    if "UHS" in headers:
        uhs_col = headers["UHS"]
    else:
        uhs_col = ws.max_column + 1
        ws.cell(row=1, column=uhs_col).value = "UHS"

    # Optional: record alpha in another column, useful for later checking
    if "UHS_alpha" in headers:
        alpha_col = headers["UHS_alpha"]
    else:
        alpha_col = ws.max_column + 1 if uhs_col <= ws.max_column else uhs_col + 1
        ws.cell(row=1, column=alpha_col).value = "UHS_alpha"

    for row in range(2, ws.max_row + 1):
        learned_recall = to_float_or_zero(ws.cell(row=row, column=recall_col).value)
        context_acc = to_float_or_zero(ws.cell(row=row, column=context_acc_col).value)

        uhs = alpha * learned_recall + (1 - alpha) * context_acc

        ws.cell(row=row, column=uhs_col).value = uhs
        ws.cell(row=row, column=uhs_col).number_format = "0.0000"

        ws.cell(row=row, column=alpha_col).value = alpha
        ws.cell(row=row, column=alpha_col).number_format = "0.00"

    wb.save(output_path)

    print(f"UHS added to sheet: {sheet_name}")
    print(f"alpha = {alpha}")
    print(f"Saved to: {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=str, required=True)
    parser.add_argument("--output", type=str, default="window_metrics_with_UHS.xlsx")
    parser.add_argument("--sheet_name", type=str, default="all_windows")
    parser.add_argument("--alpha", type=float, default=0.5)

    args = parser.parse_args()

    add_uhs_to_sheet(
        input_path=args.input,
        output_path=args.output,
        sheet_name=args.sheet_name,
        alpha=args.alpha,
    )